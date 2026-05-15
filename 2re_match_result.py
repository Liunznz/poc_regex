# import pandas as pd
# import requests
# import time
# import os
# from concurrent.futures import ThreadPoolExecutor, as_completed
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill
# import warnings
# warnings.filterwarnings('ignore', message='Unverified HTTPS request')
#
# # ====================== 配置 ======================
# API_URL = "https://222.222.44.101:18081/attack_web_atk_gm_pattern/check_pattern/"
# HEADERS = {
#     "accept": "application/json, text/plain, */*",
#     "authorization": "8bb691200ceb45bf9da3269e7ae2f40d",
#     "content-type": "application/json",
#     "origin": "https://222.222.44.101:18081",
#     "referer": "https://222.222.44.101:18081/static/attack-detection/policy-manage",
#     "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36"
# }
#
# MAX_WORKERS = 5                     # 并发线程数（可根据网络和服务端承受能力调整）
# INPUT_DIR = "poc_jihe"              # 存放原始Excel的文件夹
# OUTPUT_DIR = "test_poc"             # 输出结果文件夹
#
# COL_PATTERN = "正则字符串(直接复制)"
# COL_TEST = "所有路径(原始)"
# COL_FILTER = "漏洞中文名称(新)"
#
# RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
#
# # ====================== 验证函数（线程安全） ======================
# def verify_regex(pattern_str, test_str, row_index):
#     """调用API验证，返回 (row_index, 验证结果字典)"""
#     payload = {
#         "test_str": test_str,
#         "pattern_str": pattern_str
#     }
#     try:
#         response = requests.post(
#             url=API_URL,
#             json=payload,
#             headers=HEADERS,
#             verify=False,
#             timeout=15
#         )
#         res_data = response.json()
#         is_valid = res_data.get("is_valid", False)
#         matches = res_data.get("matches", [])
#
#         status = "成功" if is_valid and len(matches) > 0 else "失败"
#         result = {
#             "验证状态": status,
#             "匹配结果": ", ".join(matches) if matches else "",
#             "HTTP状态码": response.status_code,
#             "返回内容": response.text[:500],
#             "错误信息": ""
#         }
#     except Exception as e:
#         result = {
#             "验证状态": "异常",
#             "匹配结果": "",
#             "HTTP状态码": "请求失败",
#             "返回内容": "",
#             "错误信息": str(e)
#         }
#     return row_index, result
#
# # ====================== 处理单个Excel文件（并发） ======================
# def process_excel_file(file_path, output_dir):
#     print(f"\n📂 正在处理：{os.path.basename(file_path)}")
#     df = pd.read_excel(file_path)
#
#     # 筛选不含“未翻译”的行
#     if COL_FILTER not in df.columns:
#         print(f"⚠️ 文件缺少 '{COL_FILTER}' 列，跳过")
#         return 0, 0
#     mask = df[COL_FILTER].astype(str).str.contains("未翻译", na=False) == False
#     filtered_df = df[mask].copy()
#     if filtered_df.empty:
#         print(f"⚠️ 筛选后无有效行")
#         return 0, 0
#
#     print(f"✅ 筛选后有效行数：{len(filtered_df)}，使用 {MAX_WORKERS} 个线程并发验证")
#
#     # 准备任务列表 (原始行索引, 原始行数据, pattern, test)
#     tasks = []
#     for idx, row in filtered_df.iterrows():
#         pattern = str(row[COL_PATTERN]) if pd.notna(row[COL_PATTERN]) else ""
#         test = str(row[COL_TEST]) if pd.notna(row[COL_TEST]) else ""
#         tasks.append((idx, row, pattern, test))
#
#     # 并发执行
#     results_map = {}  # {原始索引: 验证结果字典}
#     with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
#         future_to_idx = {
#             executor.submit(verify_regex, pattern, test, idx): idx
#             for idx, row, pattern, test in tasks
#         }
#         for future in as_completed(future_to_idx):
#             idx, result = future.result()
#             results_map[idx] = result
#
#     # 按原始顺序组装结果
#     ordered_results = []
#     success_count = 0
#     for idx, row in filtered_df.iterrows():
#         verify_res = results_map[idx]
#         if verify_res["验证状态"] == "成功":
#             success_count += 1
#         combined = pd.concat([row, pd.Series(verify_res)])
#         ordered_results.append(combined)
#
#     result_df = pd.DataFrame(ordered_results)
#
#     # 保存文件并高亮失败记录
#     base_name = os.path.splitext(os.path.basename(file_path))[0]
#     out_file = os.path.join(output_dir, f"{base_name}_验证结果.xlsx")
#     result_df.to_excel(out_file, index=False, engine="openpyxl")
#
#     # 高亮失败行
#     wb = load_workbook(out_file)
#     ws = wb.active
#     status_col_letter = None
#     for col in range(1, ws.max_column + 1):
#         if ws.cell(row=1, column=col).value == "验证状态":
#             status_col_letter = ws.cell(row=1, column=col).column_letter
#             break
#     if status_col_letter:
#         for row in range(2, ws.max_row + 1):
#             cell_value = ws[f"{status_col_letter}{row}"].value
#             if cell_value in ("失败", "异常"):
#                 for col in range(1, ws.max_column + 1):
#                     ws.cell(row=row, column=col).fill = RED_FILL
#     wb.save(out_file)
#
#     print(f"📄 结果已保存：{out_file}")
#     print(f"📊 本文件成功匹配数：{success_count} / {len(filtered_df)}")
#     return success_count, len(filtered_df)
#
# # ====================== 主函数 ======================
# def main():
#     print("请确认 authorization 已修改为: 8bb691200ceb45bf9da3269e7ae2f40d")
#     confirm = input("是否继续？(y/n): ").strip().lower()
#     if confirm != 'y':
#         print("已停止，请检查 authorization 后重新运行。")
#         return
#
#     os.makedirs(OUTPUT_DIR, exist_ok=True)
#
#     pattern = "_POC解析结果.xlsx"
#     files = [f for f in os.listdir(INPUT_DIR) if f.endswith(pattern)]
#     if not files:
#         print(f"在文件夹 '{INPUT_DIR}' 中未找到任何以 '{pattern}' 结尾的文件")
#         return
#
#     print(f"📁 找到 {len(files)} 个文件：{files}")
#
#     total_success = 0
#     total_rows = 0
#
#     for file in files:
#         file_path = os.path.join(INPUT_DIR, file)
#         success, rows = process_excel_file(file_path, OUTPUT_DIR)
#         total_success += success
#         total_rows += rows
#
#     print("\n" + "="*50)
#     print(f"🎉 全部处理完成！")
#     print(f"总计成功匹配：{total_success} / {total_rows}")
#     print(f"结果保存在目录：{OUTPUT_DIR}")
#     print("="*50)
#
# if __name__ == "__main__":
#     main()


import pandas as pd
import requests
import time
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import warnings
import urllib3

# 彻底禁用所有 SSL 相关警告
warnings.filterwarnings('ignore')
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ====================== 配置 ======================
API_URL = "https://222.222.44.101:18081/attack_web_atk_gm_pattern/check_pattern/"
HEADERS = {
    "Host": "222.222.44.101:18081",
    "Sec-Ch-Ua-Platform": "\"Windows\"",
    "Authorization": "bb4f4af96d77436b9cc4be61e6e6c9d7",   # 使用你抓包得到的有效 token
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Sec-Ch-Ua": "\"Google Chrome\";v=\"147\", \"Not.A/Brand\";v=\"8\", \"Chromium\";v=\"147\"",
    "Content-Type": "application/json",
    "Sec-Ch-Ua-Mobile": "?0",
    "Origin": "https://222.222.44.101:18081",
    "Sec-Fetch-Site": "same-origin",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Dest": "empty",
    "Referer": "https://222.222.44.101:18081/static/attack-detection/policy-manage",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "zh-CN,zh;q=0.9",
    "Priority": "u=1, i",
}

MAX_WORKERS = 5
INPUT_FILE = "poc_final_fscan_nuclei_cleanedV2.xlsx"
OUTPUT_DIR = "test_poc"

COL_PATTERN = "正则字符串(直接复制)"
COL_TEST = "所有路径(原始)"
COL_FILTER = "漏洞中文名称(新)"

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# ====================== 验证函数 ======================
def verify_regex(pattern_str, test_str, row_index):
    payload = {"test_str": test_str, "pattern_str": pattern_str}
    try:
        # 关键：直接使用 verify=False，不依赖任何自定义适配器
        response = requests.post(
            API_URL,
            json=payload,
            headers=HEADERS,
            timeout=15,
            verify=False   # 禁用 SSL 证书验证
        )
        if response.status_code != 200:
            result = {
                "验证状态": "异常",
                "匹配结果": "",
                "HTTP状态码": response.status_code,
                "返回内容": response.text[:500],
                "错误信息": f"HTTP {response.status_code}"
            }
            return row_index, result

        res_data = response.json()
        is_valid = res_data.get("is_valid", False)
        matches = res_data.get("matches", [])
        status = "成功" if is_valid and len(matches) > 0 else "失败"
        result = {
            "验证状态": status,
            "匹配结果": ", ".join(matches) if matches else "",
            "HTTP状态码": response.status_code,
            "返回内容": response.text[:500],
            "错误信息": ""
        }
    except requests.exceptions.JSONDecodeError as e:
        result = {
            "验证状态": "异常",
            "匹配结果": "",
            "HTTP状态码": response.status_code if 'response' in locals() else "请求失败",
            "返回内容": response.text[:500] if 'response' in locals() else "",
            "错误信息": f"JSON 解析失败: {str(e)}"
        }
    except Exception as e:
        result = {
            "验证状态": "异常",
            "匹配结果": "",
            "HTTP状态码": "请求失败",
            "返回内容": "",
            "错误信息": str(e)
        }
    return row_index, result

# ====================== 处理 Excel 文件 ======================
def process_excel_file(file_path, output_dir):
    print(f"\n📂 正在处理：{os.path.basename(file_path)}")
    df = pd.read_excel(file_path)

    if COL_FILTER not in df.columns:
        print(f"⚠️ 文件缺少 '{COL_FILTER}' 列，跳过")
        return 0, 0

    mask = df[COL_FILTER].astype(str).str.contains("未翻译", na=False) == False
    filtered_df = df[mask].copy()
    if filtered_df.empty:
        print(f"⚠️ 筛选后无有效行")
        return 0, 0

    print(f"✅ 筛选后有效行数：{len(filtered_df)}，使用 {MAX_WORKERS} 个线程并发验证")

    tasks = []
    for idx, row in filtered_df.iterrows():
        pattern = str(row[COL_PATTERN]) if pd.notna(row[COL_PATTERN]) else ""
        test = str(row[COL_TEST]) if pd.notna(row[COL_TEST]) else ""
        tasks.append((idx, row, pattern, test))

    results_map = {}
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_idx = {
            executor.submit(verify_regex, pattern, test, idx): idx
            for idx, row, pattern, test in tasks
        }
        for future in as_completed(future_to_idx):
            idx, result = future.result()
            results_map[idx] = result

    ordered_results = []
    success_count = 0
    for idx, row in filtered_df.iterrows():
        verify_res = results_map[idx]
        if verify_res["验证状态"] == "成功":
            success_count += 1
        combined = pd.concat([row, pd.Series(verify_res)])
        ordered_results.append(combined)

    result_df = pd.DataFrame(ordered_results)

    base_name = os.path.splitext(os.path.basename(file_path))[0]
    out_file = os.path.join(output_dir, f"{base_name}_验证结果.xlsx")
    result_df.to_excel(out_file, index=False, engine="openpyxl")

    wb = load_workbook(out_file)
    ws = wb.active
    status_col_letter = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "验证状态":
            status_col_letter = ws.cell(row=1, column=col).column_letter
            break
    if status_col_letter:
        for row in range(2, ws.max_row + 1):
            cell_value = ws[f"{status_col_letter}{row}"].value
            if cell_value in ("失败", "异常"):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = RED_FILL
    wb.save(out_file)

    print(f"📄 结果已保存：{out_file}")
    print(f"📊 本文件成功匹配数：{success_count} / {len(filtered_df)}")
    return success_count, len(filtered_df)

# ====================== 主函数 ======================
def main():
    print("请确认 authorization 已替换为有效值！")
    confirm = input("是否继续？(y/n): ").strip().lower()
    if confirm != 'y':
        print("已停止，请更新 authorization 后重新运行。")
        return

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if not os.path.isfile(INPUT_FILE):
        print(f"错误：找不到文件 '{INPUT_FILE}'，请确保该文件在当前目录下。")
        return

    success, total = process_excel_file(INPUT_FILE, OUTPUT_DIR)

    print("\n" + "="*50)
    print(f"🎉 处理完成！")
    print(f"成功匹配：{success} / {total}")
    print(f"结果保存在目录：{OUTPUT_DIR}")
    print("="*50)

if __name__ == "__main__":
    main()

