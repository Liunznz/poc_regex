# import pandas as pd
# import requests
# import time
# import os
# import glob
# import threading
# from concurrent.futures import ThreadPoolExecutor, as_completed
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill
# import warnings
# warnings.filterwarnings('ignore', message='Unverified HTTPS request')
#
# # ====================== 配置 ======================
# API_URL = "https://222.222.44.101:18081/attack_web_atk_gm_pattern/"
#
# HEADERS_TEMPLATE = {
#     "accept": "application/json, text/plain, */*",
#     "authorization": "a3daa14d863f436cb355467e02628774",
#     "content-type": "application/json",
#     "origin": "https://222.222.44.101:18081",
#     "referer": "https://222.222.44.101:18081/static/attack-detection/policy-manage",
#     "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36"
# }
#
# MAX_WORKERS = 6                      # 并发线程数（可根据服务端承受能力调整）
# TARGET_OPTIONS = {
#     "1": "通用",
#     "2": "url",
#     "3": "请求体",
#     "4": "请求头"
# }
#
# RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
#
# INPUT_DIR = "test_poc"               # 存放待处理Excel的文件夹
# OUTPUT_DIR = "poc_submit"            # 输出失败记录的文件夹
#
# # 全局锁，用于保护 global_submitted_set 的原子操作
# submitted_lock = threading.Lock()
#
# # ====================== 辅助函数 ======================
# def ask_authorization():
#     print("请确认 authorization 已修改为: a3daa14d863f436cb355467e02628774")
#     confirm = input("是否继续？(y/n): ").strip().lower()
#     if confirm != 'y':
#         print("已停止，请检查 authorization 后重新运行。")
#         return False
#     return True
#
# def ask_target():
#     print("\n请选择 target 类型：")
#     for k, v in TARGET_OPTIONS.items():
#         print(f"  {k}. {v}")
#     choice = input("请输入数字 (1-4，默认2-url): ").strip() or "2"
#     target_map = {"1": "通用", "2": "url", "3": "请求体", "4": "请求头"}
#     return target_map.get(choice, "url")
#
# def send_pattern(desc, pattern_str, test_str, target, headers):
#     """单次提交，返回 (状态码, 响应文本)"""
#     payload = {
#         "unit_id": None,
#         "matches": test_str,
#         "id": None,
#         "way": "1",
#         "target": target,
#         "desc": desc,
#         "pattern_str": pattern_str,
#         "test_str": test_str,
#         "wob": "2",
#         "pattern_kws": [""]
#     }
#     try:
#         response = requests.post(url=API_URL, json=payload, headers=headers, verify=False, timeout=15)
#         return response.status_code, response.text
#     except Exception as e:
#         return "请求异常", str(e)
#
# def submit_one_row(row, target, headers, global_submitted_set):
#     """单个任务的提交逻辑，返回 (action, status_code, resp_text, row, pair_key)
#        action: 'skip' 表示全局重复跳过；'submit' 表示实际提交
#     """
#     desc = str(row["漏洞中文名称(新)"]) if pd.notna(row["漏洞中文名称(新)"]) else ""
#     pattern = str(row["正则字符串(直接复制)"]) if pd.notna(row["正则字符串(直接复制)"]) else ""
#     test = str(row["所有路径(原始)"]) if pd.notna(row["所有路径(原始)"]) else ""
#     pair_key = (desc, pattern)
#
#     # 原子检查并添加
#     with submitted_lock:
#         if pair_key in global_submitted_set:
#             return "skip", None, None, row, pair_key
#         global_submitted_set.add(pair_key)
#
#     # 实际提交
#     status_code, resp_text = send_pattern(desc, pattern, test, target, headers)
#     return "submit", status_code, resp_text, row, pair_key
#
# def process_one_file(file_path, target, headers, global_submitted_set):
#     """处理单个Excel文件（并发提交行数据），返回 (成功数, 实际提交数, 失败记录列表, 跳过数)"""
#     df = pd.read_excel(file_path)
#     if "验证状态" not in df.columns:
#         print(f"  跳过：文件缺少 '验证状态' 列")
#         return 0, 0, [], 0
#
#     success_df = df[df["验证状态"] == "成功"].copy()
#     if success_df.empty:
#         print(f"  没有验证成功的记录，跳过")
#         return 0, 0, [], 0
#
#     required_cols = ["漏洞中文名称(新)", "正则字符串(直接复制)", "所有路径(原始)"]
#     for col in required_cols:
#         if col not in success_df.columns:
#             print(f"  跳过：缺少必要列 '{col}'")
#             return 0, 0, [], 0
#
#     # 文件内去重（保留第一次出现）
#     dedup_df = success_df.drop_duplicates(subset=["漏洞中文名称(新)", "正则字符串(直接复制)"])
#     total_in_file = len(dedup_df)
#     print(f"  文件内去重后需处理 {total_in_file} 条，并发数={MAX_WORKERS}")
#
#     # 收集所有行（用于并发提交）
#     rows_list = [row for _, row in dedup_df.iterrows()]
#
#     success_count = 0
#     failed_records = []
#     skip_count = 0
#     actual_submit = 0
#
#     with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
#         # 提交所有任务
#         future_to_row = {
#             executor.submit(submit_one_row, row, target, headers, global_submitted_set): row
#             for row in rows_list
#         }
#         # 处理完成的任务
#         for future in as_completed(future_to_row):
#             action, status_code, resp_text, row, pair_key = future.result()
#             if action == "skip":
#                 skip_count += 1
#                 continue
#
#             actual_submit += 1
#             if status_code == 201:
#                 success_count += 1
#             else:
#                 fail_row = row.to_dict()
#                 fail_row["响应状态码"] = status_code
#                 fail_row["响应内容"] = resp_text[:500]
#                 failed_records.append(fail_row)
#
#     return success_count, actual_submit, failed_records, skip_count
#
# # ====================== 主流程 ======================
# def main():
#     if not ask_authorization():
#         return
#     target = ask_target()
#     print(f"已设置 target = {target}\n")
#
#     if not os.path.exists(INPUT_DIR):
#         print(f"目录不存在: {INPUT_DIR}")
#         return
#     xlsx_files = glob.glob(os.path.join(INPUT_DIR, "*.xlsx"))
#     if not xlsx_files:
#         print(f"在 {INPUT_DIR} 中没有找到任何 .xlsx 文件")
#         return
#
#     os.makedirs(OUTPUT_DIR, exist_ok=True)
#     print(f"输出目录已准备: {OUTPUT_DIR}\n")
#     print(f"找到 {len(xlsx_files)} 个文件，开始处理...\n")
#
#     headers = HEADERS_TEMPLATE.copy()
#     # 全局去重集合（跨文件）
#     global_submitted = set()
#
#     total_success_all = 0
#     total_submit_all = 0
#     total_skip_all = 0
#
#     for file_path in xlsx_files:
#         file_name = os.path.basename(file_path)
#         print(f"\n📄 处理文件: {file_name}")
#         success, submitted, failed, skipped = process_one_file(file_path, target, headers, global_submitted)
#         total_success_all += success
#         total_submit_all += submitted
#         total_skip_all += skipped
#
#         if failed:
#             fail_df = pd.DataFrame(failed)
#             out_name = file_name.replace(".xlsx", "_提交失败记录.xlsx")
#             out_path = os.path.join(OUTPUT_DIR, out_name)
#             fail_df.to_excel(out_path, index=False, engine="openpyxl")
#             # 高亮红色
#             wb = load_workbook(out_path)
#             ws = wb.active
#             for row_idx in range(2, ws.max_row + 1):
#                 for col_idx in range(1, ws.max_column + 1):
#                     ws.cell(row=row_idx, column=col_idx).fill = RED_FILL
#             wb.save(out_path)
#             print(f"  ⚠️ 失败 {len(failed)} 条，已保存至: {out_path}")
#         else:
#             print(f"  ✅ 本文件无失败记录")
#
#         print(f"  本文件实际提交 {submitted} 条，成功 {success} 条，跳过（全局重复）{skipped} 条")
#
#     print("\n" + "="*50)
#     print(f"全部处理完成！")
#     print(f"总实际提交数（去重后）: {total_submit_all}")
#     print(f"成功(201)数: {total_success_all}")
#     print(f"失败数: {total_submit_all - total_success_all}")
#     print(f"全局跳过数（重复组合）: {total_skip_all}")
#     print("="*50)
#
# if __name__ == "__main__":
#     main()
#


'''
不需要验证状态
'''
import pandas as pd
import requests
import time
import os
import glob
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import warnings
warnings.filterwarnings('ignore', message='Unverified HTTPS request')

# ====================== 配置 ======================
API_URL = "https://222.222.44.101:18081/attack_web_atk_gm_pattern/"

HEADERS_TEMPLATE = {
    "accept": "application/json, text/plain, */*",
    "authorization": "caf478373c80433d98aefdd07f3f7518",
    "content-type": "application/json",
    "origin": "https://222.222.44.101:18081",
    "referer": "https://222.222.44.101:18081/static/attack-detection/policy-manage",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36"
}

MAX_WORKERS = 6                      # 并发线程数（可根据服务端承受能力调整）
TARGET_OPTIONS = {
    "1": "通用",
    "2": "url",
    "3": "请求体",
    "4": "请求头"
}

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

INPUT_DIR = "test_poc"               # 存放待处理Excel的文件夹
OUTPUT_DIR = "poc_submit"            # 输出失败记录的文件夹

# 全局锁，用于保护 global_submitted_set 的原子操作
submitted_lock = threading.Lock()

# ====================== 辅助函数 ======================
def ask_authorization():
    print("请确认 authorization 已修改为: caf478373c80433d98aefdd07f3f7518")
    confirm = input("是否继续？(y/n): ").strip().lower()
    if confirm != 'y':
        print("已停止，请检查 authorization 后重新运行。")
        return False
    return True

def ask_target():
    print("\n请选择 target 类型：")
    for k, v in TARGET_OPTIONS.items():
        print(f"  {k}. {v}")
    choice = input("请输入数字 (1-4，默认2-url): ").strip() or "2"
    target_map = {"1": "通用", "2": "url", "3": "请求体", "4": "请求头"}
    return target_map.get(choice, "url")

def send_pattern(desc, pattern_str, test_str, target, headers):
    """单次提交，返回 (状态码, 响应文本)"""
    payload = {
        "unit_id": None,
        "matches": test_str,          # 与 test_str 相同
        "id": None,
        "way": "1",
        "target": target,
        "desc": desc,
        "pattern_str": pattern_str,
        "test_str": test_str,
        "wob": "2",
        "pattern_kws": [""]
    }
    try:
        response = requests.post(url=API_URL, json=payload, headers=headers, verify=False, timeout=15)
        return response.status_code, response.text
    except Exception as e:
        return "请求异常", str(e)

def submit_one_row(row, target, headers, global_submitted_set):
    """单个任务的提交逻辑，返回 (action, status_code, resp_text, row, pair_key)
       action: 'skip' 表示全局重复跳过；'submit' 表示实际提交
    """
    desc = str(row["漏洞中文名称(新)"]) if pd.notna(row["漏洞中文名称(新)"]) else ""
    pattern = str(row["正则字符串(直接复制)"]) if pd.notna(row["正则字符串(直接复制)"]) else ""
    test = str(row["所有路径(原始)"]) if pd.notna(row["所有路径(原始)"]) else ""

    # 检查必要字段非空
    if not desc or not pattern or not test:
        return "skip_empty", None, "必要字段为空", row, (desc, pattern)

    pair_key = (desc, pattern)

    # 原子检查并添加
    with submitted_lock:
        if pair_key in global_submitted_set:
            return "skip_duplicate", None, None, row, pair_key
        global_submitted_set.add(pair_key)

    # 实际提交
    status_code, resp_text = send_pattern(desc, pattern, test, target, headers)
    return "submit", status_code, resp_text, row, pair_key

def process_one_file(file_path, target, headers, global_submitted_set):
    """处理单个Excel文件（并发提交行数据），返回 (成功数, 实际提交数, 失败记录列表, 跳过重复数, 跳过空字段数)"""
    df = pd.read_excel(file_path)

    required_cols = ["漏洞中文名称(新)", "正则字符串(直接复制)", "所有路径(原始)"]
    for col in required_cols:
        if col not in df.columns:
            print(f"  跳过：文件缺少必要列 '{col}'")
            return 0, 0, [], 0, 0

    # 不再依赖验证状态列，直接处理所有行
    # 过滤掉三列全空的行（可选，但保留至少有一列非空的行可能无意义，这里要求三列都有值）
    # 我们将在 submit_one_row 中判断空值，并统计跳过
    rows_list = [row for _, row in df.iterrows()]

    # 文件内去重（基于 desc + pattern，保留第一次出现）
    dedup_dict = {}
    for row in rows_list:
        desc = str(row["漏洞中文名称(新)"]) if pd.notna(row["漏洞中文名称(新)"]) else ""
        pattern = str(row["正则字符串(直接复制)"]) if pd.notna(row["正则字符串(直接复制)"]) else ""
        key = (desc, pattern)
        if key not in dedup_dict:
            dedup_dict[key] = row
    dedup_rows = list(dedup_dict.values())
    total_in_file = len(dedup_rows)
    print(f"  文件内去重后需处理 {total_in_file} 条，并发数={MAX_WORKERS}")

    success_count = 0
    failed_records = []
    skip_duplicate_count = 0
    skip_empty_count = 0
    actual_submit = 0

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_row = {
            executor.submit(submit_one_row, row, target, headers, global_submitted_set): row
            for row in dedup_rows
        }
        for future in as_completed(future_to_row):
            action, status_code, resp_text, row, pair_key = future.result()
            if action == "skip_duplicate":
                skip_duplicate_count += 1
                continue
            if action == "skip_empty":
                skip_empty_count += 1
                # 可选：将空字段行也记录到失败文件
                fail_row = row.to_dict()
                fail_row["响应状态码"] = "跳过(字段为空)"
                fail_row["响应内容"] = f"desc/pattern/test 存在空值: desc={bool(row['漏洞中文名称(新)'])}, pattern={bool(row['正则字符串(直接复制)'])}, test={bool(row['所有路径(原始)'])}"
                failed_records.append(fail_row)
                continue

            actual_submit += 1
            if status_code == 201:
                success_count += 1
            else:
                fail_row = row.to_dict()
                fail_row["响应状态码"] = status_code
                fail_row["响应内容"] = resp_text[:500] if resp_text else ""
                failed_records.append(fail_row)

    return success_count, actual_submit, failed_records, skip_duplicate_count, skip_empty_count

# ====================== 主流程 ======================
def main():
    if not ask_authorization():
        return
    target = ask_target()
    print(f"已设置 target = {target}\n")

    if not os.path.exists(INPUT_DIR):
        print(f"目录不存在: {INPUT_DIR}")
        return
    xlsx_files = glob.glob(os.path.join(INPUT_DIR, "*.xlsx"))
    if not xlsx_files:
        print(f"在 {INPUT_DIR} 中没有找到任何 .xlsx 文件")
        return

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"输出目录已准备: {OUTPUT_DIR}\n")
    print(f"找到 {len(xlsx_files)} 个文件，开始处理...\n")

    headers = HEADERS_TEMPLATE.copy()
    global_submitted = set()

    total_success_all = 0
    total_submit_all = 0
    total_skip_duplicate_all = 0
    total_skip_empty_all = 0

    for file_path in xlsx_files:
        file_name = os.path.basename(file_path)
        print(f"\n📄 处理文件: {file_name}")
        success, submitted, failed, skip_dup, skip_empty = process_one_file(file_path, target, headers, global_submitted)
        total_success_all += success
        total_submit_all += submitted
        total_skip_duplicate_all += skip_dup
        total_skip_empty_all += skip_empty

        if failed:
            fail_df = pd.DataFrame(failed)
            out_name = file_name.replace(".xlsx", "_提交失败记录.xlsx")
            out_path = os.path.join(OUTPUT_DIR, out_name)
            fail_df.to_excel(out_path, index=False, engine="openpyxl")
            wb = load_workbook(out_path)
            ws = wb.active
            for row_idx in range(2, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = RED_FILL
            wb.save(out_path)
            print(f"  ⚠️ 失败/跳过 {len(failed)} 条，已保存至: {out_path}")
        else:
            print(f"  ✅ 本文件无失败/跳过记录")

        print(f"  本文件实际提交 {submitted} 条，成功 {success} 条，跳过重复（全局）{skip_dup} 条，跳过空字段 {skip_empty} 条")

    print("\n" + "="*50)
    print(f"全部处理完成！")
    print(f"总实际提交数（去重后）: {total_submit_all}")
    print(f"成功(201)数: {total_success_all}")
    print(f"失败数: {total_submit_all - total_success_all}")
    print(f"全局跳过重复数: {total_skip_duplicate_all}")
    print(f"跳过空字段行数: {total_skip_empty_all}")
    print("="*50)

if __name__ == "__main__":
    main()


















