from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string
import os

# ===================== 【用户可修改配置区】 =====================
# 原Excel文件路径
file_path = "./zhengze_quchong.xlsx"
# 目标工作表名称（留空则默认使用第一个工作表）
sheet_name = ""
# 待检查重复的列：支持列名（如"组件"）或列号（如"A"、2）
target_column = "正则字符串(直接复制)"
# 表头行号：默认第1行是表头，不参与重复检查
header_row = 1
# 重复项标记样式：True=字体标红，False=单元格填充红色
mark_font_red = True
# 新文件保存路径
new_file_path = "./duplicate_marked_result.xlsx"
# ==================================================================

def mark_duplicate_cells():
    # 校验文件是否存在
    if not os.path.exists(file_path):
        print(f"错误：未找到文件 {file_path}")
        return

    # 加载Excel工作簿
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        print(f"错误：加载Excel文件失败，{str(e)}")
        return

    # 定位目标工作表
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"错误：未找到工作表 {sheet_name}，可用工作表：{wb.sheetnames}")
            wb.close()
            return
        ws = wb[sheet_name]
    else:
        ws = wb.active
        print(f"提示：未指定工作表，默认使用第一个工作表：{ws.title}")

    # 定位目标列号
    try:
        if isinstance(target_column, int):
            # 数字列号
            col_idx = target_column
        elif target_column.isalpha():
            # 字母列号（如A、B）
            col_idx = column_index_from_string(target_column)
        else:
            # 列名：从表头行匹配
            col_idx = None
            for cell in ws[header_row]:
                if cell.value == target_column:
                    col_idx = cell.column
                    break
            if col_idx is None:
                print(f"错误：表头中未找到列名 {target_column}")
                wb.close()
                return
        col_letter = get_column_letter(col_idx)
        print(f"提示：目标列定位成功，列号：{col_idx}，列字母：{col_letter}")
    except Exception as e:
        print(f"错误：定位目标列失败，{str(e)}")
        wb.close()
        return

    # 统计列中所有值的出现次数
    value_count = {}
    # 遍历数据行（从表头下一行开始）
    for row in range(header_row + 1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=col_idx).value
        # 空值不参与重复统计
        if cell_value is not None:
            value_count[cell_value] = value_count.get(cell_value, 0) + 1

    # 筛选出重复值（出现次数>1）
    duplicate_values = {k: v for k, v in value_count.items() if v > 1}
    if not duplicate_values:
        print("提示：未检测到重复数据，无需标记")
        wb.close()
        return
    print(f"提示：检测到 {len(duplicate_values)} 组重复数据，共 {sum(duplicate_values.values())} 个重复单元格")

    # 定义重复项样式
    if mark_font_red:
        # 字体标红
        mark_style = Font(color="FF0000", bold=True)
    else:
        # 单元格填充红色（字体白色）
        from openpyxl.styles import PatternFill
        mark_style = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        font_white = Font(color="FFFFFF", bold=True)

    # 遍历单元格，标记重复项
    for row in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value in duplicate_values:
            if mark_font_red:
                cell.font = mark_style
            else:
                cell.fill = mark_style
                cell.font = font_white

    # 保存新文件
    try:
        wb.save(new_file_path)
        wb.close()
        print(f"成功：处理完成，结果已保存至 {new_file_path}")
    except Exception as e:
        print(f"错误：保存文件失败，{str(e)}")
        wb.close()

if __name__ == "__main__":
    mark_duplicate_cells()